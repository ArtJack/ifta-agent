"""Writers: cleaned CSVs, review-ready Excel, gov-portal CSV."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ifta.calc import IftaReturn, StateLine
from ifta.models import CleanData

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Cleaned data files
# ---------------------------------------------------------------------------


def write_cleaned_csvs(data: CleanData, out_dir: Path) -> tuple[Path, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    miles_path = out_dir / "cleaned_miles.csv"
    fuel_path = out_dir / "cleaned_fuel.csv"

    with miles_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["truck_id", "state", "miles"])
        for mileage_record in sorted(data.miles, key=lambda row: (row.truck_id, row.state)):
            w.writerow(
                [
                    mileage_record.truck_id,
                    mileage_record.state,
                    round(mileage_record.miles, 2),
                ]
            )

    with fuel_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["truck_id", "state", "gallons", "tax_paid"])
        for fuel_record in sorted(data.fuel, key=lambda row: (row.truck_id, row.state)):
            w.writerow(
                [
                    fuel_record.truck_id,
                    fuel_record.state,
                    round(fuel_record.gallons, 3),
                    round(fuel_record.tax_paid, 2),
                ]
            )

    return miles_path, fuel_path


# ---------------------------------------------------------------------------
# Owner-review Excel (multi-sheet audit workbook)
# ---------------------------------------------------------------------------

MONEY_FMT = "$#,##0.00;[Red]($#,##0.00)"


def _set_header(cell: object) -> None:
    c = cast(object, cell)
    c.font = HEADER_FONT  # type: ignore[attr-defined]
    c.fill = HEADER_FILL  # type: ignore[attr-defined]
    c.alignment = CENTER  # type: ignore[attr-defined]


def _autosize(ws: Worksheet, max_cols: int, width: int = 14) -> None:
    for c in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def _build_cover_sheet(
    ws: Worksheet,
    *,
    ret: IftaReturn,
    client_name: str | None,
    review_summary: str | None,
) -> None:
    """First sheet — at-a-glance view for the owner."""
    ws.title = "Cover"
    ws["A1"] = f"IFTA Quarterly Return — {ret.quarter}"
    ws["A1"].font = Font(bold=True, size=16)

    rows: list[tuple[str, object, str | None]] = [
        ("Carrier", client_name or "(client not set)", None),
        ("Quarter", ret.quarter, None),
        ("Fuel type", ret.fuel.title(), None),
        ("Fleet trucks", ", ".join(t.truck_id for t in ret.trucks), None),
        ("Fleet miles", round(ret.fleet_miles, 0), "#,##0"),
        ("Fleet gallons", round(ret.fleet_gallons, 2), "#,##0.00"),
        ("Fleet MPG", round(ret.fleet_mpg, 2), "0.00"),
        ("Total tax due", round(ret.total_tax_due, 2), MONEY_FMT),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        a = ws.cell(row=i, column=1, value=label)
        a.font = TOTAL_FONT
        b = ws.cell(row=i, column=2, value=value)
        if fmt:
            b.number_format = fmt

    if ret.rate_fallback_used:
        ws.cell(row=len(rows) + 4, column=1, value="⚠ RATE FALLBACK").font = Font(
            bold=True, color="C00000"
        )
        ws.cell(
            row=len(rows) + 4, column=2, value=ret.rate_warning or "Confirm rates."
        ).font = Font(color="C00000")

    if review_summary:
        ws.cell(row=len(rows) + 6, column=1, value="Agent summary").font = TOTAL_FONT
        ws.cell(row=len(rows) + 7, column=1, value=review_summary).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.merge_cells(
            start_row=len(rows) + 7, start_column=1, end_row=len(rows) + 11, end_column=6
        )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42


def _build_per_truck_sheet(ws: Worksheet, ret: IftaReturn, data: CleanData) -> None:
    """Per-truck blocks in David's `IFTA 2025 ACTIVED.xlsx` layout.

    For each truck, a 5-column block: TRUCK | STATE | MILES | GALLONS | MPG.
    All trucks shown side-by-side, then a right-most summary block with the
    fleet jurisdiction summary (CDTFA-style columns).
    """
    ws.title = "Per-Truck Worksheet"
    ws["A1"] = f"IFTA {ret.quarter} — Per-Truck Worksheet"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Fleet MPG"
    ws["A3"].font = TOTAL_FONT
    ws["B3"] = round(ret.fleet_mpg, 4)
    ws["A4"] = "Total Tax Due"
    ws["A4"].font = TOTAL_FONT
    ws["B4"] = round(ret.total_tax_due, 2)
    ws["B4"].number_format = MONEY_FMT

    miles_by_tk_st: dict[tuple[str, str], float] = {}
    for mileage_record in data.miles:
        miles_by_tk_st[(mileage_record.truck_id, mileage_record.state)] = (
            miles_by_tk_st.get((mileage_record.truck_id, mileage_record.state), 0.0)
            + mileage_record.miles
        )
    gallons_by_tk_st: dict[tuple[str, str], float] = {}
    for fuel_record in data.fuel:
        gallons_by_tk_st[(fuel_record.truck_id, fuel_record.state)] = (
            gallons_by_tk_st.get((fuel_record.truck_id, fuel_record.state), 0.0)
            + fuel_record.gallons
        )

    # Union of states each truck visits, sorted alphabetically — matches
    # the layout in David's workbook (consistent state column per block).
    states = [line.state for line in ret.lines if not line.is_surcharge]
    header_row = 6
    data_start = header_row + 1

    col = 1
    for t in ret.trucks:
        _set_header(ws.cell(row=header_row, column=col, value=t.truck_id))
        for offset, label in enumerate(("STATE", "MILES", "GALLONS", "MPG"), start=1):
            _set_header(ws.cell(row=header_row, column=col + offset, value=label))

        for i, s in enumerate(states):
            row_idx = data_start + i
            m = miles_by_tk_st.get((t.truck_id, s), 0.0)
            g = gallons_by_tk_st.get((t.truck_id, s), 0.0)
            ws.cell(row=row_idx, column=col, value=t.truck_id)
            ws.cell(row=row_idx, column=col + 1, value=s)
            ws.cell(row=row_idx, column=col + 2, value=round(m, 2) if m else None)
            ws.cell(row=row_idx, column=col + 3, value=round(g, 3) if g else None)
            if m and g:
                ws.cell(row=row_idx, column=col + 4, value=round(m / g, 2))

        total_row = data_start + len(states)
        ws.cell(row=total_row, column=col + 1, value="TOTAL").font = TOTAL_FONT
        ws.cell(row=total_row, column=col + 2, value=round(t.miles, 2)).font = TOTAL_FONT
        ws.cell(row=total_row, column=col + 3, value=round(t.gallons, 3)).font = TOTAL_FONT
        ws.cell(row=total_row, column=col + 4, value=round(t.mpg, 4)).font = TOTAL_FONT
        col += 5  # next block (one blank col between would also work)

    # Summary block on the right
    sum_col = col + 1
    sum_headers = [
        "Jurisdiction",
        "Surcharge",
        "Total Miles",
        "Taxable Miles",
        "MPG",
        "Taxable Gal",
        "Tax Paid Gal",
        "Net Taxable Gal",
        "Rate",
        "Tax",
    ]
    for i, h in enumerate(sum_headers):
        _set_header(ws.cell(row=header_row, column=sum_col + i, value=h))

    for i, line in enumerate(ret.lines):
        row_idx = data_start + i
        ws.cell(row=row_idx, column=sum_col, value=line.state)
        ws.cell(row=row_idx, column=sum_col + 1, value="Surcharge" if line.is_surcharge else "")
        ws.cell(row=row_idx, column=sum_col + 2, value=int(round(line.miles)))
        ws.cell(row=row_idx, column=sum_col + 3, value=int(round(line.miles)))
        ws.cell(
            row=row_idx,
            column=sum_col + 4,
            value=0.0 if line.is_surcharge else round(ret.fleet_mpg, 2),
        )
        ws.cell(row=row_idx, column=sum_col + 5, value=int(round(line.taxable_gal)))
        ws.cell(row=row_idx, column=sum_col + 6, value=int(round(line.tax_paid_gal)))
        ws.cell(row=row_idx, column=sum_col + 7, value=int(round(line.net_taxable_gal)))
        ws.cell(row=row_idx, column=sum_col + 8, value=round(line.rate, 4))
        c = ws.cell(row=row_idx, column=sum_col + 9, value=round(line.tax_due, 2))
        c.number_format = MONEY_FMT

    sum_total_row = data_start + len(ret.lines)
    ws.cell(row=sum_total_row, column=sum_col, value="TOTAL").font = TOTAL_FONT
    ws.cell(
        row=sum_total_row, column=sum_col + 2, value=int(round(ret.fleet_miles))
    ).font = TOTAL_FONT
    ws.cell(
        row=sum_total_row, column=sum_col + 3, value=int(round(ret.fleet_miles))
    ).font = TOTAL_FONT
    total_cell = ws.cell(row=sum_total_row, column=sum_col + 9, value=round(ret.total_tax_due, 2))
    total_cell.font = TOTAL_FONT
    total_cell.number_format = MONEY_FMT

    _autosize(ws, sum_col + len(sum_headers))


def _build_jurisdiction_sheet(ws: Worksheet, ret: IftaReturn) -> None:
    """The exact same table that gets uploaded to the portal."""
    ws.title = "Jurisdiction Summary"
    ws["A1"] = f"Jurisdiction Summary — {ret.quarter}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "This is the data uploaded to the gov portal."
    ws["A2"].font = Font(italic=True, color="595959")

    headers = [
        "Jurisdiction",
        "Surcharge",
        "Total Miles",
        "Taxable Miles",
        "MPG",
        "Taxable Gal",
        "Tax Paid Gal",
        "Net Taxable Gal",
        "Rate",
        "Tax",
        "Interest",
        "Total",
    ]
    header_row = 4
    for i, h in enumerate(headers):
        _set_header(ws.cell(row=header_row, column=1 + i, value=h))

    for i, line in enumerate(ret.lines):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=line.state)
        ws.cell(row=r, column=2, value="Surcharge" if line.is_surcharge else "")
        ws.cell(row=r, column=3, value=int(round(line.miles)))
        ws.cell(row=r, column=4, value=int(round(line.miles)))
        ws.cell(row=r, column=5, value=0.0 if line.is_surcharge else round(ret.fleet_mpg, 2))
        ws.cell(row=r, column=6, value=int(round(line.taxable_gal)))
        ws.cell(row=r, column=7, value=int(round(line.tax_paid_gal)))
        ws.cell(row=r, column=8, value=int(round(line.net_taxable_gal)))
        ws.cell(row=r, column=9, value=round(line.rate, 4))
        ws.cell(row=r, column=10, value=round(line.tax_due, 2)).number_format = MONEY_FMT
        ws.cell(row=r, column=11, value=0.0).number_format = MONEY_FMT
        ws.cell(row=r, column=12, value=round(line.tax_due, 2)).number_format = MONEY_FMT

    total_row = header_row + 1 + len(ret.lines)
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=3, value=int(round(ret.fleet_miles))).font = TOTAL_FONT
    ws.cell(row=total_row, column=4, value=int(round(ret.fleet_miles))).font = TOTAL_FONT
    c10 = ws.cell(row=total_row, column=10, value=round(ret.total_tax_due, 2))
    c10.font = TOTAL_FONT
    c10.number_format = MONEY_FMT
    ws.cell(row=total_row, column=11, value=0.0).number_format = MONEY_FMT
    c12 = ws.cell(row=total_row, column=12, value=round(ret.total_tax_due, 2))
    c12.font = TOTAL_FONT
    c12.number_format = MONEY_FMT

    _autosize(ws, len(headers))


def _build_review_sheet(
    ws: Worksheet,
    *,
    review: dict[str, object] | None,
    metrics: dict[str, object] | None,
) -> None:
    """Agent narrative + run metrics."""
    ws.title = "Agent Review"
    ws["A1"] = "Agent Review"
    ws["A1"].font = Font(bold=True, size=14)
    r = 3
    if review is None:
        ws.cell(row=r, column=1, value="Agent review was not run for this quarter.")
        return

    summary = review.get("summary") if isinstance(review, dict) else None
    if summary:
        ws.cell(row=r, column=1, value="Summary").font = TOTAL_FONT
        r += 1
        ws.cell(row=r, column=1, value=str(summary)).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.merge_cells(start_row=r, start_column=1, end_row=r + 3, end_column=6)
        r += 5

    for section_label, key in (
        ("Issues", "issues"),
        ("Filing reminders", "filing_reminders"),
        ("Next steps", "next_steps"),
    ):
        items = review.get(key) if isinstance(review, dict) else None
        if not items:
            continue
        from ifta.agent import format_review_item

        ws.cell(row=r, column=1, value=section_label).font = TOTAL_FONT
        r += 1
        review_items = items if isinstance(items, list) else [items]
        for item in review_items:
            ws.cell(
                row=r,
                column=1,
                value=f"• {format_review_item(cast(Any, item))}",
            ).alignment = Alignment(wrap_text=True)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            r += 1
        r += 1

    if metrics:
        ws.cell(row=r, column=1, value="Run metrics").font = TOTAL_FONT
        r += 1
        for label, value in metrics.items():
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=value)
            r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def write_owner_review_xlsx(
    ret: IftaReturn,
    data: CleanData,
    out_path: Path,
    *,
    client_name: str | None = None,
    review: dict[str, object] | None = None,
    metrics: dict[str, object] | None = None,
) -> Path:
    """Customer-facing multi-sheet workbook.

    Sheets:
        1. Cover                — carrier, totals, agent verdict
        2. Per-Truck Worksheet  — David's per-truck-block format + summary
        3. Jurisdiction Summary — same data as the portal CSV
        4. Agent Review         — narrative + run metrics
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # The first sheet `Sheet` is created automatically — use it for Cover.
    cover = cast(Worksheet, wb.active)
    review_summary = str(review.get("summary")) if review and isinstance(review, dict) else None
    _build_cover_sheet(cover, ret=ret, client_name=client_name, review_summary=review_summary)
    _build_per_truck_sheet(wb.create_sheet(), ret, data)
    _build_jurisdiction_sheet(wb.create_sheet(), ret)
    _build_review_sheet(wb.create_sheet(), review=review, metrics=metrics)
    wb.save(out_path)
    return out_path


# Backwards-compatible alias for older callers/tests.
write_review_xlsx = write_owner_review_xlsx


# ---------------------------------------------------------------------------
# Gov-portal CSV (per-state filing rows)
# ---------------------------------------------------------------------------


def _portal_fuel_label(fuel: str, portal: str) -> str:
    normalized_fuel = fuel.lower().replace(" ", "_")
    normalized_portal = portal.lower().strip()
    if normalized_portal == "cdtfa":
        return {"diesel": "2. Diesel", "gasoline": "1. Gasoline"}.get(normalized_fuel, fuel)
    return {"diesel": "Diesel", "gasoline": "Gasoline"}.get(normalized_fuel, fuel.title())


def write_portal_csv(ret: IftaReturn, out_path: Path, *, portal: str = "generic") -> Path:
    """IFTA Quarterly Return per-jurisdiction worksheet CSV.

    The column layout is a human-review worksheet inspired by the CDTFA
    Jurisdiction Summary. It is not guaranteed to be a direct-upload schema
    for every base-state portal; each state may use different fuel codes,
    column names, and import rules.

      Jurisdiction | Surcharge | Fuel Type | Total Miles | Taxable Miles |
      MPG | Taxable Gal | Tax Paid Gal | Net Taxable Gal | Rate | Tax |
      Interest | Total

    Gallons are rounded to whole integers (portal convention). Surcharge
    rows have miles=0, MPG=0, and use the surcharge rate.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fuel_label = _portal_fuel_label(ret.fuel, portal)

    def jur_name(line) -> str:
        # CDTFA uses full state names; for now keep the 2-letter code
        # (portals accept either; readers can mentally expand)
        return line.state

    with out_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        if ret.rate_fallback_used:
            w.writerow(["WARNING", ret.rate_warning or "Fallback rates were used."])
            w.writerow(
                [
                    "DO_NOT_FILE",
                    "Confirm the current-quarter IFTA rate matrix before portal submission.",
                ]
            )
            w.writerow([])
        w.writerow(
            [
                "Jurisdiction",
                "Surcharge",
                "Fuel Type",
                "Total Miles",
                "Taxable Miles",
                "MPG",
                "Taxable Gal",
                "Tax Paid Gal",
                "Net Taxable Gal",
                "Rate",
                "Tax",
                "Interest",
                "Total",
            ]
        )
        for line in ret.lines:
            mpg = "0.00" if line.is_surcharge else f"{ret.fleet_mpg:.2f}"
            w.writerow(
                [
                    jur_name(line),
                    "Surcharge" if line.is_surcharge else "",
                    fuel_label,
                    int(round(line.miles)),
                    int(round(line.miles)),  # taxable miles = total (no exempt-miles claim)
                    mpg,
                    int(round(line.taxable_gal)),
                    int(round(line.tax_paid_gal)),
                    int(round(line.net_taxable_gal)),
                    f"{line.rate:.4f}",
                    f"{line.tax_due:.2f}",
                    "0.00",
                    f"{line.tax_due:.2f}",
                ]
            )
        total_miles = int(round(ret.fleet_miles))
        w.writerow(
            [
                "TOTAL",
                "",
                "",
                total_miles,
                total_miles,
                "",
                "",
                "",
                "",
                "",
                f"{ret.total_tax_due:.2f}",
                "0.00",
                f"{ret.total_tax_due:.2f}",
            ]
        )
    return out_path


# ---------------------------------------------------------------------------
# Per-truck Jurisdiction Summary (Menshikov-CDTFA style, one file per truck)
# ---------------------------------------------------------------------------


def _truck_filename(truck_id: str) -> str:
    """Filesystem-safe truck filename. e.g. 'T1' -> 'truck_T1.xlsx'."""
    safe = "".join(c if c.isalnum() else "_" for c in truck_id).strip("_") or "unknown"
    return f"truck_{safe}.xlsx"


RECONCILIATION_NOTE = (
    "These numbers are this truck's share of the carrier's IFTA filing. "
    "Each row shows the truck's miles and gallons in that state and its "
    "proportional contribution to the per-state IFTA tax (computed using "
    "fleet MPG and per-state rates). Sums across all trucks in the fleet "
    "equal the totals filed at the gov portal — i.e. one truck's per-state "
    "credit or liability may not match the fleet line for that state."
)


def write_truck_filing_xlsx(
    truck_id: str,
    lines: list[StateLine],
    *,
    fleet_mpg: float,
    quarter: str,
    client_name: str,
    fuel: str,
    out_path: Path,
    period_begin: str | None = None,
    period_end: str | None = None,
    driver: str | None = None,
    card_number: str | None = None,
) -> Path:
    """Single-sheet Menshikov-CDTFA-style Jurisdiction Summary for one truck.

    Layout mirrors the CDTFA filing PDFs (header info + Jurisdiction Summary
    table). Math uses fleet MPG so per-truck totals reconcile back to the
    fleet filing across all trucks.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = f"Truck {truck_id}"

    # ----- Title block -----
    ws["A1"] = f"{client_name} — IFTA {quarter}"
    ws["A1"].font = Font(bold=True, size=14)
    subtitle_parts = [f"Truck {truck_id}"]
    if driver:
        subtitle_parts.append(f"Driver: {driver}")
    if card_number:
        subtitle_parts.append(f"Card #: {card_number}")
    ws["A2"] = " · ".join(subtitle_parts) + " — Per-Jurisdiction Detail"
    ws["A2"].font = Font(italic=True, color="595959")

    # ----- Reconciliation note (P0a) — explains why per-truck ≠ portal -----
    note_cell = ws.cell(row=3, column=1, value=RECONCILIATION_NOTE)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    note_cell.font = Font(italic=True, color="404040", size=10)
    note_cell.fill = PatternFill("solid", fgColor="FFF8E1")  # soft amber
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=13)
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 18

    # ----- Header info rows -----
    truck_total_miles = sum(ln.miles for ln in lines if not ln.is_surcharge)
    truck_total_taxable_gal = sum(ln.taxable_gal for ln in lines if not ln.is_surcharge)
    truck_total_tax_paid_gal = sum(ln.tax_paid_gal for ln in lines if not ln.is_surcharge)
    truck_total_net_gal = sum(ln.net_taxable_gal for ln in lines if not ln.is_surcharge)
    truck_net_tax = sum(ln.tax_due for ln in lines)

    period_str = f"{period_begin} – {period_end}" if period_begin and period_end else "—"
    info_rows: list[tuple[str, object, str | None]] = [
        ("Quarter", quarter, None),
        ("Fuel type", fuel.title(), None),
        ("Period", period_str, None),
        ("Truck total miles", int(round(truck_total_miles)), "#,##0"),
        ("Gallons purchased", round(truck_total_tax_paid_gal, 2), "#,##0.00"),
        ("Fleet MPG (used for tax calc)", round(fleet_mpg, 2), "0.00"),
        ("Truck's share of net tax", round(truck_net_tax, 2), MONEY_FMT),
    ]
    info_start = 7  # below the merged reconciliation note
    for i, (label, value, fmt) in enumerate(info_rows, start=info_start):
        a = ws.cell(row=i, column=1, value=label)
        a.font = TOTAL_FONT
        b = ws.cell(row=i, column=2, value=value)
        if fmt:
            b.number_format = fmt

    # ----- Jurisdiction Summary table -----
    table_start = info_start + len(info_rows) + 1
    ws.cell(row=table_start, column=1, value="Jurisdiction Summary").font = Font(bold=True, size=12)
    header_row = table_start + 1
    headers = [
        "Jurisdiction",
        "Surcharge",
        "Fuel Type",
        "Total Miles",
        "Taxable Miles",
        "MPG",
        "Taxable Gal",
        "Tax Paid Gal",
        "Net Taxable Gal",
        "Rate",
        "Tax",
        "Interest",
        "Total",
    ]
    for i, h in enumerate(headers):
        _set_header(ws.cell(row=header_row, column=1 + i, value=h))

    fuel_label = "Diesel" if fuel.lower() == "diesel" else fuel.title()

    for i, ln in enumerate(lines):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=ln.state)
        ws.cell(row=r, column=2, value="Surcharge" if ln.is_surcharge else "")
        ws.cell(row=r, column=3, value=fuel_label)
        ws.cell(row=r, column=4, value=int(round(ln.miles)))
        ws.cell(row=r, column=5, value=int(round(ln.miles)))
        ws.cell(row=r, column=6, value=0.0 if ln.is_surcharge else round(fleet_mpg, 2))
        ws.cell(row=r, column=7, value=int(round(ln.taxable_gal)))
        ws.cell(row=r, column=8, value=int(round(ln.tax_paid_gal)))
        ws.cell(row=r, column=9, value=int(round(ln.net_taxable_gal)))
        ws.cell(row=r, column=10, value=round(ln.rate, 4))
        ws.cell(row=r, column=11, value=round(ln.tax_due, 2)).number_format = MONEY_FMT
        ws.cell(row=r, column=12, value=0.0).number_format = MONEY_FMT
        ws.cell(row=r, column=13, value=round(ln.tax_due, 2)).number_format = MONEY_FMT

    # ----- TOTAL row (P0c — fill the gallon totals too) -----
    total_row = header_row + 1 + len(lines)
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=4, value=int(round(truck_total_miles))).font = TOTAL_FONT
    ws.cell(row=total_row, column=5, value=int(round(truck_total_miles))).font = TOTAL_FONT
    ws.cell(row=total_row, column=7, value=int(round(truck_total_taxable_gal))).font = TOTAL_FONT
    ws.cell(row=total_row, column=8, value=int(round(truck_total_tax_paid_gal))).font = TOTAL_FONT
    ws.cell(row=total_row, column=9, value=int(round(truck_total_net_gal))).font = TOTAL_FONT
    c11 = ws.cell(row=total_row, column=11, value=round(truck_net_tax, 2))
    c11.font = TOTAL_FONT
    c11.number_format = MONEY_FMT
    ws.cell(row=total_row, column=12, value=0.0).number_format = MONEY_FMT
    c13 = ws.cell(row=total_row, column=13, value=round(truck_net_tax, 2))
    c13.font = TOTAL_FONT
    c13.number_format = MONEY_FMT

    # ----- Column widths (P0b — wider so headers don't clip) -----
    # Per-column widths sized to the longer of the header label or the
    # typical data shape. "Net Taxable Gal" is 15 chars → width 17.
    ws.column_dimensions["A"].width = 30  # truck-header labels live in col A
    ws.column_dimensions["B"].width = 12
    column_widths = {
        3: 11,  # Fuel Type
        4: 13,  # Total Miles
        5: 14,  # Taxable Miles
        6: 7,  # MPG
        7: 12,  # Taxable Gal
        8: 13,  # Tax Paid Gal
        9: 17,  # Net Taxable Gal
        10: 8,  # Rate
        11: 11,  # Tax
        12: 10,  # Interest
        13: 11,  # Total
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)
    return out_path


def write_per_truck_filings(
    per_truck_lines: dict[str, list[StateLine]],
    *,
    fleet_mpg: float,
    quarter: str,
    client_name: str,
    fuel: str,
    out_dir: Path,
    period_begin: str | None = None,
    period_end: str | None = None,
    data: CleanData | None = None,
) -> list[Path]:
    """Write one Excel per truck. Returns the list of written paths.

    - If period_begin / period_end aren't supplied, they're derived from the
      quarter (e.g. "Q2-2026" → "April 1, 2026" / "June 30, 2026").
    - If `data` is supplied, each truck's driver name and fuel card number
      are read from it and rendered on the file's title block.
    """
    from ifta.client import quarter_dates

    if period_begin is None and period_end is None:
        period_begin, period_end = quarter_dates(quarter)

    out_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for truck_id, lines in per_truck_lines.items():
        p = out_dir / _truck_filename(truck_id)
        write_truck_filing_xlsx(
            truck_id,
            lines,
            fleet_mpg=fleet_mpg,
            quarter=quarter,
            client_name=client_name,
            fuel=fuel,
            out_path=p,
            period_begin=period_begin,
            period_end=period_end,
            driver=data.driver(truck_id) if data else None,
            card_number=data.card(truck_id) if data else None,
        )
        paths.append(p)
    return paths
