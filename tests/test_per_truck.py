"""Per-truck Jurisdiction Summary tests.

The key invariants:
1. Sum of per-truck miles in a state == fleet line miles for that state
2. Sum of per-truck tax (across trucks) approximately equals fleet total tax
3. Surcharge lines appear per truck only when that truck has miles in that state
4. Per-truck Excel files actually get written, one per truck
"""

from pathlib import Path

from conftest import rates_or_skip

from ifta.calc import compute_per_truck_lines, compute_return
from ifta.ingest import ingest_folder
from ifta.report import write_per_truck_filings

ROOT = Path(__file__).resolve().parents[1]


def _load(quarter: str):
    data = ingest_folder(ROOT / "inbox" / quarter)
    # Q2-2026 rates aren't committed as a cache file, so this needs network.
    # rates_or_skip skips (not fails) when the box is offline.
    rates = rates_or_skip(quarter)
    ret = compute_return(data, rates)
    per_truck = compute_per_truck_lines(data, ret, rates)
    return data, ret, rates, per_truck


# ---------------------------------------------------------------------------
# Invariant: per-truck miles per state sum to fleet miles per state
# ---------------------------------------------------------------------------


def test_per_truck_miles_reconcile_to_fleet_miles_q2_2026() -> None:
    _, ret, _, per_truck = _load("Q2-2026")

    fleet_miles_by_state: dict[str, float] = {}
    for ln in ret.lines:
        if ln.is_surcharge:
            continue
        fleet_miles_by_state[ln.state] = fleet_miles_by_state.get(ln.state, 0.0) + ln.miles

    truck_miles_by_state: dict[str, float] = {}
    for lines in per_truck.values():
        for ln in lines:
            if ln.is_surcharge:
                continue
            truck_miles_by_state[ln.state] = truck_miles_by_state.get(ln.state, 0.0) + ln.miles

    for state, fleet_v in fleet_miles_by_state.items():
        assert truck_miles_by_state.get(state, 0.0) == fleet_v, (
            state,
            fleet_v,
            truck_miles_by_state.get(state),
        )


# ---------------------------------------------------------------------------
# Invariant: per-truck tax sums approximately equal fleet tax
# ---------------------------------------------------------------------------


def test_per_truck_tax_reconciles_to_fleet_total_q2_2026() -> None:
    _, ret, _, per_truck = _load("Q2-2026")
    truck_total = sum(ln.tax_due for lines in per_truck.values() for ln in lines)
    # Tolerance: per-state integer rounding can drift cents across trucks.
    # 1 cent per state per truck is typical; cap the slack generously.
    assert abs(truck_total - ret.total_tax_due) < 5.0, (truck_total, ret.total_tax_due)


# ---------------------------------------------------------------------------
# Surcharge per truck only when miles in that state
# ---------------------------------------------------------------------------


def test_surcharge_only_when_truck_has_miles_in_state() -> None:
    _, _, _, per_truck = _load("Q2-2026")
    for truck_id, lines in per_truck.items():
        miles_by_state = {ln.state: ln.miles for ln in lines if not ln.is_surcharge}
        for ln in lines:
            if ln.is_surcharge:
                truck_miles_in_state = miles_by_state.get(ln.state, 0.0)
                assert truck_miles_in_state > 0, (
                    f"Surcharge for {ln.state} on truck {truck_id} but no miles there"
                )


# ---------------------------------------------------------------------------
# Writer: one xlsx file per truck, no missing or extra files
# ---------------------------------------------------------------------------


def test_write_per_truck_filings_creates_one_file_per_truck(tmp_path: Path) -> None:
    _, ret, _, per_truck = _load("Q2-2026")
    paths = write_per_truck_filings(
        per_truck,
        fleet_mpg=ret.fleet_mpg,
        quarter=ret.quarter,
        client_name="TEST CARRIER",
        fuel=ret.fuel,
        out_dir=tmp_path,
    )
    assert len(paths) == len(per_truck)
    assert all(p.exists() for p in paths)
    truck_ids = {p.stem.replace("truck_", "") for p in paths}
    assert truck_ids == set(per_truck.keys())


def test_truck_filing_xlsx_has_expected_structure(tmp_path: Path) -> None:
    """Sanity-check one written file actually contains the expected rows."""
    import openpyxl

    _, ret, _, per_truck = _load("Q2-2026")
    truck_id, lines = next(iter(per_truck.items()))
    paths = write_per_truck_filings(
        {truck_id: lines},
        fleet_mpg=ret.fleet_mpg,
        quarter=ret.quarter,
        client_name="TEST CARRIER",
        fuel=ret.fuel,
        out_dir=tmp_path,
    )
    wb = openpyxl.load_workbook(paths[0])
    ws = wb.active
    # Header text should include client + quarter
    assert "TEST CARRIER" in str(ws["A1"].value)
    assert ret.quarter in str(ws["A1"].value)
    assert str(truck_id) in str(ws["A2"].value)

    # Find the Jurisdiction Summary header row
    cells = [(row, ws.cell(row=row, column=1).value) for row in range(1, 30)]
    summary_row = next(r for r, v in cells if v == "Jurisdiction Summary")
    headers = [ws.cell(row=summary_row + 1, column=c).value for c in range(1, 14)]
    assert headers[0] == "Jurisdiction"
    assert "Tax" in headers
